"""
ویوهای سفارشی پنل ادمین امور مالی
"""
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, Count
from django.utils import timezone
from datetime import timedelta, datetime
import jdatetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from .models import Invoice


@staff_member_required
def current_revenue_report_view(request):
    """گزارش درآمد جاری - فقط نمودارهای هفتگی و ماهانه"""
    
    # محاسبه درآمد هفته جاری (شروع از شنبه)
    today = timezone.now()
    # پیدا کردن شنبه این هفته (weekday: شنبه=5)
    days_since_saturday = (today.weekday() + 2) % 7
    week_start = (today - timedelta(days=days_since_saturday)).replace(hour=0, minute=0, second=0, microsecond=0)
    week_end = week_start + timedelta(days=7)
    
    weekly_data = []
    week_labels = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنج‌شنبه', 'جمعه']
    
    for i in range(7):
        day_start = week_start + timedelta(days=i)
        day_end = day_start + timedelta(days=1)
        
        daily_revenue = Invoice.objects.filter(
            status='paid',
            paid_at__gte=day_start,
            paid_at__lt=day_end
        ).aggregate(total=Sum('total'))['total'] or 0
        
        weekly_data.append(float(daily_revenue))
    
    # محاسبه درآمد ماه جاری (شمسی)
    now_jalali = jdatetime.datetime.now()
    month_start_jalali = now_jalali.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # تبدیل به میلادی
    month_start_gregorian = month_start_jalali.togregorian()
    month_start = timezone.make_aware(datetime.combine(month_start_gregorian, datetime.min.time()))
    
    # محاسبه تعداد روزهای ماه شمسی
    if now_jalali.month <= 6:
        days_in_month = 31
    elif now_jalali.month <= 11:
        days_in_month = 30
    else:
        # بهمن - بررسی کبیسه
        days_in_month = 30 if jdatetime.j_days_in_month[12](now_jalali.year) == 30 else 29
    
    monthly_data = []
    monthly_labels = []
    
    for day in range(1, days_in_month + 1):
        day_jalali = now_jalali.replace(day=day, hour=0, minute=0, second=0, microsecond=0)
        day_start = timezone.make_aware(datetime.combine(day_jalali.togregorian(), datetime.min.time()))
        day_end = day_start + timedelta(days=1)
        
        daily_revenue = Invoice.objects.filter(
            status='paid',
            paid_at__gte=day_start,
            paid_at__lt=day_end
        ).aggregate(total=Sum('total'))['total'] or 0
        
        monthly_data.append(float(daily_revenue))
        monthly_labels.append(str(day))
    
    
    # محاسبه مجموع درآمد هفته و ماه
    week_total = sum(weekly_data)
    month_total = sum(monthly_data)
    
    context = {
        'weekly_labels': week_labels,
        'weekly_data': weekly_data,
        'week_total': week_total,
        'monthly_labels': monthly_labels,
        'monthly_data': monthly_data,
        'month_total': month_total,
        'current_month_name': now_jalali.strftime('%B %Y'),
    }
    
    return render(request, 'admin/finance/current_revenue_report.html', context)


@staff_member_required
def date_range_revenue_report_view(request):
    """گزارش درآمد بازه تاریخی با نمودار و خروجی اکسل"""
    
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    export_excel = request.GET.get('export') == 'excel'
    
    custom_revenue = None
    custom_invoice_count = None
    daily_data = []
    daily_labels = []
    invoices_list = []
    
    if date_from and date_to:
        try:
            # تبدیل تاریخ شمسی به میلادی
            from_parts = date_from.split('-')
            to_parts = date_to.split('-')
            
            from_jalali = jdatetime.date(int(from_parts[0]), int(from_parts[1]), int(from_parts[2]))
            to_jalali = jdatetime.date(int(to_parts[0]), int(to_parts[1]), int(to_parts[2]))
            
            from_gregorian = from_jalali.togregorian()
            to_gregorian = to_jalali.togregorian()
            
            custom_date_from = timezone.make_aware(datetime.combine(from_gregorian, datetime.min.time()))
            custom_date_to = timezone.make_aware(datetime.combine(to_gregorian, datetime.max.time()))
            
            # دریافت فاکتورها
            invoices = Invoice.objects.filter(
                status='paid',
                paid_at__gte=custom_date_from,
                paid_at__lte=custom_date_to
            ).order_by('paid_at')
            
            custom_stats = invoices.aggregate(
                total=Sum('total'),
                count=Count('id')
            )
            
            custom_revenue = custom_stats['total'] or 0
            custom_invoice_count = custom_stats['count'] or 0
            
            # محاسبه درآمد روزانه برای نمودار
            current_date = from_gregorian
            while current_date <= to_gregorian:
                day_start = timezone.make_aware(datetime.combine(current_date, datetime.min.time()))
                day_end = day_start + timedelta(days=1)
                
                daily_revenue = Invoice.objects.filter(
                    status='paid',
                    paid_at__gte=day_start,
                    paid_at__lt=day_end
                ).aggregate(total=Sum('total'))['total'] or 0
                
                # تبدیل به شمسی برای نمایش
                jalali_date = jdatetime.date.fromgregorian(date=current_date)
                daily_labels.append(jalali_date.strftime('%Y/%m/%d'))
                daily_data.append(float(daily_revenue))
                
                current_date += timedelta(days=1)
            
            # آماده‌سازی لیست فاکتورها
            for invoice in invoices:
                jalali_paid_at = jdatetime.datetime.fromgregorian(datetime=invoice.paid_at.replace(tzinfo=None))
                invoices_list.append({
                    'invoice_number': invoice.invoice_number,
                    'buyer_name': invoice.buyer_name,
                    'total': invoice.total,
                    'paid_at': jalali_paid_at.strftime('%Y/%m/%d %H:%M'),
                })
            
            # اگر درخواست خروجی اکسل است
            if export_excel:
                return generate_excel_report(invoices_list, date_from, date_to, custom_revenue, custom_invoice_count)
                
        except (ValueError, IndexError) as e:
            pass
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'custom_revenue': custom_revenue,
        'custom_invoice_count': custom_invoice_count,
        'daily_labels': daily_labels,
        'daily_data': daily_data,
        'invoices_list': invoices_list,
    }
    
    return render(request, 'admin/finance/daterange_revenue_report.html', context)


def generate_excel_report(invoices_list, date_from, date_to, total_revenue, invoice_count):
    """تولید فایل اکسل گزارش درآمد"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'گزارش درآمد'
    
    # تنظیمات راست به چپ
    ws.sheet_view.rightToLeft = True
    
    # هدر گزارش
    ws['A1'] = f'گزارش درآمد از {date_from} تا {date_to}'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f'تعداد فاکتورها: {invoice_count}'
    ws['A2'].font = Font(bold=True)
    ws['A3'] = f'مجموع درآمد: {total_revenue:,.0f} تومان'
    ws['A3'].font = Font(bold=True)
    
    # هدر جدول
    headers = ['شماره فاکتور', 'نام خریدار', 'مبلغ (تومان)', 'تاریخ پرداخت']
    header_fill = PatternFill(start_color='417690', end_color='417690', fill_type='solid')
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # داده‌های فاکتورها
    for row_idx, invoice in enumerate(invoices_list, start=6):
        ws.cell(row=row_idx, column=1, value=invoice['invoice_number'])
        ws.cell(row=row_idx, column=2, value=invoice['buyer_name'])
        ws.cell(row=row_idx, column=3, value=float(invoice['total']))
        ws.cell(row=row_idx, column=4, value=invoice['paid_at'])
    
    # تنظیم عرض ستون‌ها
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    # ذخیره در حافظه و ارسال
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="revenue_report_{date_from}_to_{date_to}.xlsx"'
    wb.save(response)
    
    return response
